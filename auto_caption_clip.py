import argparse
import os
import re
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from PIL import Image

import open_clip
import torch



from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook as XLWorkbook


FIELD_ORDER = [
    "Body Portion",
    "Race",
    "Pose",
    "Dress Types",
    "Location",
    "Props",  # optional, multi-label
]

IMAGE_EXTS = (".png", ".jpg", ".jpeg", ".webp", ".bmp")


def natural_index_from_name(name: str) -> int:
    m = re.search(r"(\d+)", name)
    return int(m.group(1)) if m else 0


def find_images(images_dir: str) -> List[str]:
    files = [
        os.path.join(images_dir, f)
        for f in os.listdir(images_dir)
        if os.path.splitext(f)[1].lower() in IMAGE_EXTS
    ]
    files.sort(key=lambda p: natural_index_from_name(os.path.basename(p)))
    return files


def read_guide_options(path: str) -> Dict[str, List[str]]:
    options: Dict[str, List[str]] = {k: [] for k in FIELD_ORDER}
    df = pd.read_excel(path)
    # Map columns by case-insensitive matching
    colmap: Dict[str, str] = {}
    for c in df.columns:
        key = str(c).strip()
        low = key.lower()
        for target in FIELD_ORDER:
            if low == target.lower():
                colmap[target] = key
                break
    for target in FIELD_ORDER:
        if target in colmap:
            vals = (
                df[colmap[target]]
                .dropna()
                .astype(str)
                .map(lambda s: s.strip())
                .replace({"": pd.NA})
                .dropna()
                .unique()
                .tolist()
            )
            options[target] = list(sorted(vals))
    return options


def ensure_workbook(path: str):
    if os.path.exists(path):
        try:
            return load_workbook(path)
        except Exception:
            pass
    wb = XLWorkbook()
    ws = wb.active
    ws.title = "Sheet1"
    wb.save(path)
    return wb


def write_caption(path: str, row_index: int, caption: str):
    wb = ensure_workbook(path)
    ws = wb[wb.sheetnames[0]]
    ws.cell(row=row_index, column=1).value = caption
    wb.save(path)


def write_artifact(path: str, row_index: int, filename: str, score: float):
    wb = ensure_workbook(path)
    name = "artifacts"
    ws = wb[name] if name in wb.sheetnames else wb.create_sheet(name)
    target_row = ws.max_row + 1 if ws.max_row else 1
    ws.cell(row=target_row, column=1).value = row_index
    ws.cell(row=target_row, column=2).value = filename
    ws.cell(row=target_row, column=3).value = float(score)
    wb.save(path)


def normalize(t: torch.Tensor) -> torch.Tensor:
    return t / t.norm(dim=-1, keepdim=True).clamp(min=1e-12)


def build_prompts(field: str, options: List[str]) -> List[str]:
    # Light field-specific prompting
    if field == "Body Portion":
        return [f"a person, {opt}" for opt in options]
    if field == "Race":
        return [f"a person of {opt}" for opt in options]
    if field == "Pose":
        return [f"a person in pose {opt}" for opt in options]
    if field == "Dress Types":
        return [f"a person wearing {opt}" for opt in options]
    if field == "Location":
        return [f"outdoor scene at {opt}" if any(w in opt.lower() for w in ["beach","street","park","forest"]) else f"indoor scene at {opt}" for opt in options]
    if field == "Props":
        return [f"a person with {opt}" for opt in options]
    return options


def encode_texts(tokenizer, model, device, texts: List[str], batch_size: int = 64) -> torch.Tensor:
    embs: List[torch.Tensor] = []
    for i in range(0, len(texts), batch_size):
        batch = texts[i : i + batch_size]
        tokens = tokenizer(batch).to(device)
        with torch.no_grad(), torch.cuda.amp.autocast(enabled=False):
            feats = model.encode_text(tokens)
            feats = normalize(feats.float())
        embs.append(feats)
    return torch.cat(embs, dim=0)


def encode_images(preprocess, model, device, paths: List[str], batch_size: int = 16) -> torch.Tensor:
    embs: List[torch.Tensor] = []
    for i in range(0, len(paths), batch_size):
        batch = []
        for p in paths[i : i + batch_size]:
            img = Image.open(p).convert("RGB")
            batch.append(preprocess(img))
        imgs = torch.stack(batch).to(device)
        with torch.no_grad(), torch.cuda.amp.autocast(enabled=False):
            feats = model.encode_image(imgs)
            feats = normalize(feats.float())
        embs.append(feats.cpu())
    return torch.cat(embs, dim=0)


def choose_single(sim: np.ndarray, options: List[str]) -> Optional[str]:
    if sim.size == 0:
        return None
    idx = int(sim.argmax())
    return options[idx]


def choose_props(sim: np.ndarray, options: List[str], max_props: int, floor: float, band: float) -> List[str]:
    if sim.size == 0:
        return []
    best = float(sim.max())
    picks: List[Tuple[int, float]] = []
    for i, s in enumerate(sim.tolist()):
        if s >= max(floor, best - band):
            picks.append((i, s))
    picks.sort(key=lambda t: t[1], reverse=True)
    picks = picks[:max_props]
    return [options[i] for i, _ in picks]


def artifact_score(image_feat: torch.Tensor, tokenizer, model, device) -> float:
    prompts = [
        "photo with body deformation, distorted body, extra limbs, missing torso, fused body",
        "photo of a normal human body"
    ]
    txt = encode_texts(tokenizer, model, device, prompts)
    sim = (image_feat @ txt.T).cpu().numpy()[0]
    # positive if deformed > normal
    return float(sim[0] - sim[1])


def main():
    parser = argparse.ArgumentParser(description="Automatic captioning with CLIP slot-filling")
    parser.add_argument("--images_dir", required=True)
    parser.add_argument("--guide", required=True)
    parser.add_argument("--output", required=True)
    parser.add_argument("--model", default="ViT-B-32")
    parser.add_argument("--pretrained", default="openai")
    parser.add_argument("--batch_size", type=int, default=16)
    parser.add_argument("--limit", type=int, default=0, help="Process only first N images if >0")
    parser.add_argument("--props_max", type=int, default=5)
    parser.add_argument("--props_floor", type=float, default=0.24)
    parser.add_argument("--props_band", type=float, default=0.02, help="Within best-score band to include")
    parser.add_argument("--artifact", action="store_true", help="Compute simple CLIP artifact score and log if suspicious")
    parser.add_argument("--artifact_threshold", type=float, default=0.05, help="Flag if deform-normal > threshold")
    args = parser.parse_args()

    device = "cuda" if torch.cuda.is_available() else "cpu"
    model, _, preprocess = open_clip.create_model_and_transforms(args.model, pretrained=args.pretrained)
    tokenizer = open_clip.get_tokenizer(args.model)
    model = model.to(device)
    model.eval()

    images = find_images(args.images_dir)
    if args.limit and args.limit > 0:
        images = images[: args.limit]
    if not images:
        raise SystemExit(f"No images in {args.images_dir}")

    options = read_guide_options(args.guide)

    # Pre-embed options per field
    field_text_embs: Dict[str, torch.Tensor] = {}
    field_text_opts: Dict[str, List[str]] = {}
    for field in FIELD_ORDER:
        opts = options.get(field, [])
        opts = [o for o in opts if o]
        if not opts:
            continue
        prompts = build_prompts(field, opts)
        embs = encode_texts(tokenizer, model, device, prompts)
        field_text_embs[field] = embs
        field_text_opts[field] = opts

    # Encode all images
    img_feats = encode_images(preprocess, model, device, images, batch_size=args.batch_size)

    # For each image, pick options
    for idx, img_path in enumerate(images):
        feat = img_feats[idx : idx + 1]  # (1, d)
        caption_parts: List[str] = []
        for field in FIELD_ORDER:
            opts = field_text_opts.get(field, [])
            if not opts:
                continue
            txt = field_text_embs[field]
            sim = (feat @ txt.T).cpu().numpy()[0]
            if field == "Props":
                picks = choose_props(sim, opts, args.props_max, args.props_floor, args.props_band)
                if picks:
                    caption_parts.append(", ".join(picks))
            else:
                sel = choose_single(sim, opts)
                if sel:
                    caption_parts.append(sel)

        caption = ", ".join([p for p in caption_parts if p]).rstrip(" ,.")
        row_index = idx + 1
        write_caption(args.output, row_index, caption)

        if args.artifact:
            score = artifact_score(feat, tokenizer, model, device)
            if score > args.artifact_threshold:
                write_artifact(args.output, row_index, os.path.basename(img_path), score)

    print(f"Wrote {len(images)} captions to {args.output}")


if __name__ == "__main__":
    main()
