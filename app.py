import argparse
import os
import re
import sys
from datetime import datetime
from typing import Dict, List, Optional

import pandas as pd
from PIL import Image, ImageTk
import tkinter as tk
from tkinter import ttk, messagebox

try:
    from openpyxl import load_workbook
    from openpyxl.workbook.workbook import Workbook
except Exception:  # pragma: no cover - handled at runtime
    load_workbook = None  # type: ignore
    Workbook = None  # type: ignore


FIELD_ORDER = [
    "Body Portion",
    "Race",
    "Pose",
    "Dress Types",
    "Location",
    "Props",  # optional
]

IMAGE_EXTS = (".png", ".jpg", ".jpeg", ".webp", ".bmp")


def natural_index_from_name(name: str) -> int:
    m = re.search(r"(\d+)", name)
    return int(m.group(1)) if m else 0


def find_images(images_dir: str) -> List[str]:
    files = [
        os.path.join(images_dir, f)
        for f in os.listdir(images_dir)
        if os.path.splitext(f)[1].lower() in IMAGE_EXTS
    ]
    files.sort(key=lambda p: natural_index_from_name(os.path.basename(p)))
    return files


def read_guide_options(path: str) -> Dict[str, List[str]]:
    options: Dict[str, List[str]] = {k: [] for k in FIELD_ORDER}
    if not os.path.exists(path):
        return options

    # Try to read the first sheet
    df = pd.read_excel(path)
    # normalize columns for matching
    colmap: Dict[str, str] = {}
    for c in df.columns:
        key = str(c).strip()
        low = key.lower()
        for target in FIELD_ORDER:
            if low == target.lower():
                colmap[target] = key
                break
    for target in FIELD_ORDER:
        if target in colmap:
            vals = (
                df[colmap[target]]
                .dropna()
                .astype(str)
                .map(lambda s: s.strip())
                .replace({"": pd.NA})
                .dropna()
                .unique()
                .tolist()
            )
            options[target] = list(sorted(vals))
    return options


def ensure_workbook(path: str):
    from openpyxl import Workbook as XLWorkbook  # local import

    if os.path.exists(path):
        try:
            wb = load_workbook(path)
            return wb
        except Exception:
            raise
    wb = XLWorkbook()
    ws = wb.active
    ws.title = "Sheet1"
    wb.save(path)
    return wb


def write_caption(path: str, row_index: int, caption: str):
    wb = ensure_workbook(path)
    ws = wb[wb.sheetnames[0]]
    ws.cell(row=row_index, column=1).value = caption
    wb.save(path)


def write_artifact(path: str, row_index: int, filename: str, note: Optional[str], flagged: bool):
    wb = ensure_workbook(path)
    sheet_name = "artifacts"
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

    # find if entry exists for this row_index or filename
    found_row = None
    for r in range(1, ws.max_row + 1):
        ridx = ws.cell(row=r, column=1).value
        fname = ws.cell(row=r, column=2).value
        if ridx == row_index or fname == filename:
            found_row = r
            break
    target_row = found_row or (ws.max_row + 1 if ws.max_row else 1)
    ws.cell(row=target_row, column=1).value = row_index
    ws.cell(row=target_row, column=2).value = filename
    ws.cell(row=target_row, column=3).value = (note or "").strip()
    ws.cell(row=target_row, column=4).value = "flagged" if flagged else "unflagged"
    ws.cell(row=target_row, column=5).value = datetime.now().isoformat(timespec="seconds")
    wb.save(path)


def read_existing_caption(path: str, row_index: int) -> Optional[str]:
    if not os.path.exists(path):
        return None
    try:
        wb = load_workbook(path)
        ws = wb[wb.sheetnames[0]]
        val = ws.cell(row=row_index, column=1).value
        return str(val) if val is not None else None
    except Exception:
        return None


class CaptionApp:
    def __init__(self, root: tk.Tk, images_dir: str, guide_path: str, output_path: str):
        self.root = root
        self.images_dir = images_dir
        self.guide_path = guide_path
        self.output_path = output_path

        self.images = find_images(self.images_dir)
        if not self.images:
            messagebox.showerror("No images", f"No images found in: {self.images_dir}")
            sys.exit(1)

        self.options = read_guide_options(self.guide_path)
        self.index = 0
        self.flag_artifact = False

        self._build_ui()
        self._load_image()
        self._bind_keys()

    def _build_ui(self):
        self.root.title("Caption Helper")
        self.root.geometry("1200x800")

        # Top controls
        top = ttk.Frame(self.root)
        top.pack(side=tk.TOP, fill=tk.X, padx=8, pady=8)

        self.info_var = tk.StringVar()
        ttk.Label(top, textvariable=self.info_var).pack(side=tk.LEFT)

        ttk.Button(top, text="Prev (Left)", command=self.prev_image).pack(side=tk.LEFT, padx=4)
        ttk.Button(top, text="Save & Next (Enter)", command=self.save_and_next).pack(side=tk.LEFT, padx=4)
        ttk.Button(top, text="Save (Ctrl+S)", command=self.save_only).pack(side=tk.LEFT, padx=4)

        self.artifact_var = tk.BooleanVar()
        artifact_chk = ttk.Checkbutton(top, text="Body Artifact (Alt+A)", variable=self.artifact_var, command=self._artifact_toggled)
        artifact_chk.pack(side=tk.LEFT, padx=10)

        # Main area: left image, right fields
        main = ttk.Frame(self.root)
        main.pack(fill=tk.BOTH, expand=True)

        self.canvas = tk.Canvas(main, width=800, height=700, bg="#222")
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 8))

        right = ttk.Frame(main)
        right.pack(side=tk.LEFT, fill=tk.Y)

        self.field_vars: Dict[str, tk.StringVar] = {}
        self.combos: Dict[str, ttk.Combobox] = {}
        for field in FIELD_ORDER:
            frm = ttk.Frame(right)
            frm.pack(fill=tk.X, pady=4)
            ttk.Label(frm, text=field + (" (optional)" if field == "Props" else "")).pack(anchor=tk.W)
            var = tk.StringVar()
            self.field_vars[field] = var
            combo = ttk.Combobox(frm, textvariable=var, values=self.options.get(field, []))
            combo.set("")
            combo.pack(fill=tk.X)
            combo['state'] = 'normal'  # allow typing new values
            self.combos[field] = combo

        note_frame = ttk.Frame(right)
        note_frame.pack(fill=tk.X, pady=(12, 4))
        ttk.Label(note_frame, text="Artifact note (optional)").pack(anchor=tk.W)
        self.note_entry = ttk.Entry(note_frame)
        self.note_entry.pack(fill=tk.X)

        # Caption preview
        prev_frame = ttk.Frame(right)
        prev_frame.pack(fill=tk.X, pady=(12, 4))
        ttk.Label(prev_frame, text="Caption preview").pack(anchor=tk.W)
        self.caption_var = tk.StringVar()
        self.caption_label = ttk.Label(prev_frame, textvariable=self.caption_var, wraplength=350)
        self.caption_label.pack(fill=tk.X)

        # Update preview when fields change
        for var in self.field_vars.values():
            var.trace_add('write', lambda *args: self._update_preview())

    def _bind_keys(self):
        self.root.bind('<Left>', lambda e: self.prev_image())
        self.root.bind('<Right>', lambda e: self.save_and_next())
        self.root.bind('<Return>', lambda e: self.save_and_next())
        self.root.bind('<Control-s>', lambda e: self.save_only())
        self.root.bind('<Alt-a>', lambda e: self._toggle_artifact())

    def _artifact_toggled(self):
        self.flag_artifact = bool(self.artifact_var.get())

    def _toggle_artifact(self):
        self.artifact_var.set(not self.artifact_var.get())
        self._artifact_toggled()

    def _compose_caption(self) -> str:
        parts: List[str] = []
        for field in FIELD_ORDER:
            val = (self.field_vars[field].get() or "").strip()
            if not val:
                continue
            parts.append(val)
        caption = ", ".join([p for p in parts if p])
        # Ensure caption doesn't end with comma or period (it shouldn't by construction)
        caption = caption.rstrip(" ,.")
        return caption

    def _update_preview(self):
        self.caption_var.set(self._compose_caption())

    def _current_row_index(self) -> int:
        # row 1 corresponds to first image (index 0)
        return self.index + 1

    def _load_image(self):
        img_path = self.images[self.index]
        self.info_var.set(f"{os.path.basename(img_path)}  |  {self.index + 1} / {len(self.images)}")
        # show image fitted into canvas
        try:
            img = Image.open(img_path)
            cw = int(self.canvas.winfo_width() or 800)
            ch = int(self.canvas.winfo_height() or 700)
            img.thumbnail((cw, ch))
            self.tk_img = ImageTk.PhotoImage(img)
            self.canvas.delete("all")
            self.canvas.create_image(cw // 2, ch // 2, image=self.tk_img)
        except Exception as e:
            self.canvas.delete("all")
            self.canvas.create_text(10, 10, anchor='nw', fill='white', text=f"Error loading image: {e}")
        # If there is an existing caption saved and fields are empty, show it
        existing = read_existing_caption(self.output_path, self._current_row_index())
        if existing and not self._compose_caption():
            self.caption_var.set(existing)
        else:
            self._update_preview()

    def _save_caption_and_artifact(self) -> bool:
        caption = self._compose_caption()
        row_idx = self._current_row_index()
        try:
            # Preserve existing if new caption is empty
            if not caption:
                existing = read_existing_caption(self.output_path, row_idx)
                if existing:
                    # nothing to do
                    pass
                else:
                    write_caption(self.output_path, row_idx, caption)
            else:
                write_caption(self.output_path, row_idx, caption)
            # artifact optional
            if self.flag_artifact:
                note = self.note_entry.get().strip()
                write_artifact(
                    self.output_path,
                    row_idx,
                    os.path.basename(self.images[self.index]),
                    note=note,
                    flagged=True,
                )
            return True
        except PermissionError:
            messagebox.showerror("Save failed", "Unable to save. Close the Excel file if it's open and try again.")
            return False
        except Exception as e:
            messagebox.showerror("Save failed", f"Error saving caption: {e}")
            return False

    def prev_image(self):
        if self.index > 0:
            self.index -= 1
            self._load_image()

    def save_only(self):
        self._save_caption_and_artifact()

    def save_and_next(self):
        if self._save_caption_and_artifact():
            if self.index < len(self.images) - 1:
                self.index += 1
                self._load_image()
            else:
                messagebox.showinfo("Done", "Reached the last image. All saved.")


def main():
    parser = argparse.ArgumentParser(description="Captioning helper")
    parser.add_argument("--images_dir", default=os.path.join(os.path.dirname(__file__), "..", "1_indian_woman"))
    parser.add_argument("--guide", default=os.path.join(os.path.dirname(__file__), "..", "Caption_guide_reg.xlsx"))
    parser.add_argument("--output", default=os.path.join(os.path.dirname(__file__), "..", "captions_prompts.xlsx"))
    args = parser.parse_args()

    images_dir = os.path.abspath(args.images_dir)
    guide_path = os.path.abspath(args.guide)
    output_path = os.path.abspath(args.output)

    root = tk.Tk()
    app = CaptionApp(root, images_dir, guide_path, output_path)
    root.mainloop()


if __name__ == "__main__":
    main()
